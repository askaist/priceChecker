import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class Util:

    # takes in a list of url from wassermans products and returns a list of the prices for those products
    @staticmethod
    def getPriceFromWassermans(urlList):
        priceList = []

        # Initialize the WebDriver
        driver = webdriver.Chrome()

        # Navigate to the website
        for url in urlList:
            driver.get(url)
            time.sleep(1)
            # element_locator = (By.ID, "myElement")
            wait = WebDriverWait(driver, 10)  # 10 seconds timeout
            price_element = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'prices')))
            price_text = price_element.text
            priceList.append(price_text)
            print("cost: " + price_text)

        driver.quit()

        return priceList

    # takes in a list of prices and displays them downwards in an excel sheet starting
    # starting with the cell with the coordinates [startRow, startCol]
    @staticmethod
    def addListToPrices(list, startRow, startCol):
        # Load the existing workbook
        workbook = load_workbook('priceCheckerWorkBook.xlsx')

        # Select the active sheet
        worksheet = workbook['prices']

        for i in range(0, len(list)):
            char = get_column_letter(startCol)
            cell = char + str(startRow + i)
            worksheet[cell] = list[i]
            print(list[i] + " added to cell " + cell)

        # Save the workbook
        workbook.save('priceCheckerWorkBook.xlsx')

        print("data saved")

    @staticmethod
    def getPricesFromWalmart():
        driver = webdriver.Chrome()
        driver.get("https://www.walmart.com/")
        time.sleep(100)

    @staticmethod
    def saveUrlListFromExcel(startRow, startCol):
        url = ""
        urlList = []

        # Load the existing workbook
        workbook = load_workbook('priceCheckerWorkBook.xlsx')

        # Select the active sheet
        worksheet = workbook['urls']

        i = 0
        while True:
            char = get_column_letter(startCol)
            cell_value = worksheet[char + str(startRow + i)].value

            if cell_value is None:
                break  # Exit the loop when the cell is empty (None)

            urlList.append(cell_value)
            i += 1
            print(cell_value)

        return urlList

    @staticmethod
    def getPricesFromAarons(urlList):
        priceList = []

        # Initialize the WebDriver
        driver = webdriver.Chrome()

        driver.get("https://www.aronskissenafarms.com/")
        wait = WebDriverWait(driver, 10)
        cityInput = wait.until(EC.visibility_of_element_located((By.ID, 'filter-areas-input')))

        cityInput.send_keys('queens')
        cityInput.send_keys(Keys.RETURN)
        button = wait.until(EC.visibility_of_element_located((By.XPATH, '//button[contains(@ng-click, '
                                                                        '"chooseAreaCtrl.chooseArea.submit")]')))
        button.click()

        # Navigate to the website
        for url in urlList:
            driver.get(url)
            time.sleep(1)

            # element_locator = (By.ID, "myElement")
            wait = WebDriverWait(driver, 10)  # 10 seconds timeout

            # Find all elements with either class 'price' or 'sale-price'
            price_elements = driver.find_elements(By.CSS_SELECTOR, '.price, .sale-price')


            # Get the first element
            price_element = wait.until(EC.visibility_of(price_elements[2]))
            price_text = price_element.text
            priceList.append(price_text)
            print("cost:", price_text)


        driver.quit()

        return priceList

    @staticmethod
    def getPricesFromSeasons(urlList):
        priceList = []

        # Initialize the WebDriver
        driver = webdriver.Chrome()

        # Navigate to the website
        for url in urlList:
            driver.get(url)
            time.sleep(1)
            # element_locator = (By.ID, "myElement")
            wait = WebDriverWait(driver, 10)  # 10 seconds timeout
            price_element = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'product-price')))
            price_text = price_element.text
            priceList.append(price_text)
            print("cost: " + price_text)

        driver.quit()

        return priceList
